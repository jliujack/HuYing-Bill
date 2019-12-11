
import tkinter
import tkinter.messagebox
import os
import difflib
import re
import yaml
import time

from enum import Enum
from openpyxl import load_workbook
from handleBill import copyWordFile,assist,docx_enhanced,util
import handleBill.assist as assist

PENDING = 0
PROCESSING = 1
FUFILLED = 2
ERROR = 3

current_path = os.path.dirname(os.path.realpath(__file__))
f = open(current_path+"\\config.yml", encoding="utf-8")
globelConfig = yaml.load(f, Loader=yaml.FullLoader)
f.close()

CONFIG_COL, WORK_COL = util.CONFIG_COL, util.WORK_COL

class HANDEL_METHOD(Enum):
  NORMAL = 0

class Bill:
  '账单处理'
  process = PENDING
  workMap = {}
  notContain = []
  tempDir = []
  targetDir = []
  config = []
  partner = {}
 
  def __init__(self, window, excelFile, targetPath, templateWordPath, defaultTempFile, printLog):
    self.window = window
    self.excelFile = excelFile
    self.targetPath = targetPath
    self.templateWordPath = templateWordPath
    self.defaultTempFile = defaultTempFile
    self.printLog = printLog

  def displayexcelPath(self):
    print ("hello:" ,self.excelFile.get())

  def handleBill(self):

    self.getTemplateWordsFilePath()
    self.getConfig()
    self.handleExcel()
    self.handleWords()

  def getConfig(self):
    print("getConfig")
    try:
      wb = load_workbook("./config/配置文件.xlsx")
      sheet = wb["fileConfig"]
      self.configList = [[i.value for i in item] for item in list(sheet)[1:]] #caseName, fileName, handleType, contractNum, versionNum, handleMethod
      nameList = [ [i.value for i in item] for item in list(wb["nameMap"])[1:]]
      self.nameMap = {}
      for item in nameList:
        self.nameMap[item[0]] = {"shortName": item[1], "fullName": item[2], "cname": item[0]}
      self.partner = {"shortName": "", "fullName": "", "cname": ""}
      if globelConfig['partner'] in self.nameMap:
        self.partner = self.nameMap[globelConfig['partner']]
    except :
      raise assist.CustomError('请检查配置文件是否有问题！')

  def handleExcel(self):
    try:
      print("handle excel")
      wb = load_workbook(self.excelFile.get())
      sheet = wb["导出结果"]
      workItems = list(sheet)
      workItems = workItems[1:]
      workMap = {}
      self.workMap = workMap
      for item in workItems:
        key = item[WORK_COL.caseName.value].value
        if key not in workMap:
          workMap[key] = [[i.value for i in item]]
        else:
          workMap[key].append([i.value for i in item])

      print('keys:',list(workMap.keys()))
      print(len(list(workItems)))
    except :
      raise assist.CustomError('请检查工作时间文件是否有问题！')
    

  def getTemplateWordsFilePath(self):
    try:
      print("getTemplateWords")
      dirs = os.walk(self.templateWordPath.get())
      locDir = next(dirs)
      self.tempDir = {"abPath": locDir[0], "files": [i for i in locDir[2] if re.search(r'docx?$',i)]}
    except :
       raise assist.CustomError('请检查模板文件是否有问题！')
    

  def handleWords(self):
    self.notContain = []
    self.normalContain = []
    startTime = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
    self.printLog("开始时间：" + startTime)
    for config in self.configList:
      if config[CONFIG_COL.caseName.value] not in self.workMap:
        self.notContain.append(config[CONFIG_COL.caseName.value])
        continue
      if config[CONFIG_COL.handleMethod.value] == HANDEL_METHOD.NORMAL.value:
        self.handleNormalWord(config)
        self.normalContain.append(config[CONFIG_COL.caseName.value])
    if len(self.notContain) != 0:
      text = "[" + ", ".join(self.notContain) + "]没有工作记录"
      tkinter.messagebox.showinfo(title='warining', message=text)
      self.printLog(text)
    text = "["+ ", ".join(self.normalContain) + "]处理成功"
    self.printLog(text)
    endTime = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
    self.printLog("结束时间：" + endTime)

  def handleNormalWord(self, config):
    try:
      tempPath = self.defaultTempFile
      fileName = ""
      date = config[CONFIG_COL.specifiedDate.value]
      if date == "":
        date = time.strftime("%Y-%m-%d", time.localtime())
      if config[CONFIG_COL.handleType.value] == "update":
        fileName = difflib.get_close_matches( config[CONFIG_COL.fileName.value],self.tempDir["files"],1,0.2)
        tempPath = self.tempDir["abPath"] + "/" + fileName[0]

      date = str(date)[0:10]
      targetFileName = config[CONFIG_COL.fileName.value] + "_[" + str(config[CONFIG_COL.contractNum .value]) + "_0001." + \
        str(config[CONFIG_COL.versionNum.value]) + "]_" + date + ".docx"
      targetPath = self.targetPath.get() + "/" + targetFileName
      copyWordFile.copyToTargetPath(tempPath, targetPath)
      caseWorkItems = assist.handelArray(self.workMap[config[CONFIG_COL.caseName.value]])
      billStatistics, total = self.getStatistics(caseWorkItems, config)
      docx_enhanced.replaceNormalWord(targetPath, caseWorkItems, billStatistics, total, date, config)
    except Exception as e:
      print(e)
    

  def getStatistics(self, caseWorkItems, config):
    try:
      ret = []
      total = 0
      lawyerRate = {}
      rateList = config[CONFIG_COL.lawyerRate.value].split(';')
      rateItems = {rate.split(':')[1]:rate.split(':')[0] for rate in rateList}
      for key in rateItems:
        obj = {i:key for i in rateItems[key].split(',')}
        lawyerRate.update(obj)
    
      summary = {}
      keys = []
      for item in caseWorkItems:
        if lawyerRate[item[2]] not in summary:
          summary[lawyerRate[item[2]]] = {"fullNames":[self.nameMap[item[2]]["fullName"]],"time": item[3], "rate": lawyerRate[item[2]]}
        else:
          summary[lawyerRate[item[2]]]["time"] += item[3]
          summary[lawyerRate[item[2]]]["fullNames"].append(self.nameMap[item[2]]["fullName"])
        item[2] = self.nameMap[item[2]]["shortName"]
      for key in summary:
        summary[key]["fullNames"] = list(set(summary[key]["fullNames"]))
        flag = False
        for i in range(len(summary[key]["fullNames"])):
          if summary[key]["fullNames"][i] == self.partner["fullName"]:
            summary[key]["fullNames"][i],summary[key]["fullNames"][0] = summary[key]["fullNames"][0],summary[key]["fullNames"][i]
            keys = [key] + keys
            flag = True
            break
        if not flag:
          if len(summary[key]["fullNames"]) == 1:
            summary[key]["fullNames"] = summary[key]["fullNames"][0]
          else:
            summary[key]["fullNames"] = "/".join(summary[key]["fullNames"])
          keys = keys + [key]
      for i in range(len(keys)):
        data = summary[keys[i]]
        ret.append([data["fullNames"],data["time"],data["rate"],data["time"] * int(data["rate"])])
        total += data["time"] * float(data["rate"])
      return ret,total
    except Exception as e:
      print(e)
    






